
from config import UPLOAD_FOLDER
import openpyxl
from app.models import Drasdocument, Drasrevision, Drascommentsheet, Drascomment, Splitofworks, Unitmodel, Disciplinedras, Tagdiscipline
from datetime import datetime
from flask import flash, abort
from app import db
from random import randint
from datetime import timedelta


def date_parse(field):
    if isinstance(field, datetime):
        #print('We got a valid Date!')
        return field
    try:
        return datetime.strptime(field,'%d/%m/%y')
    except:
        #print('Date Parse ERROR --------------- GOT:', field)
        return None

def check_labels(item):
    # 
    # Check File Labels
    # 
    #  
    item = item.cs_file
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item)
    csSheet = csFile.active
    
    # Duplicate key (label in excel) for check purpose
    header_labels_dict = {
        'Reference' : csSheet['C8'].value,
        'Date' : csSheet['D8'].value,

        'Reference' : csSheet['G8'].value,
        'Date' : csSheet['G9'].value,
        'Material requisition': csSheet['G10'].value,
        'Vendor Name' : csSheet['G11'].value,

        'Rev.' : csSheet['K9'].value,
        'Description': csSheet['K10'].value,
        'Issued by (Contractor Discipline)': csSheet['K11'].value
    }

    for key, value in header_labels_dict.items():
        #print(key, value)
        if key != value:
            #flash(('Header Label ' + key + ' Not Found, Check your DRAS Template!'), category='warning')
            abort(400,('La Label ' + key + ' non è nella posizione corretta, controllare le colonne del DRAS.'))
    
    table_label_dict = {
        'Pos.' : csSheet['B14'].value,
        
        'Status' : csSheet['G16'].value,

        'Date' : csSheet['L15'].value,

    }
    
    return True

def update_data_from_cs(item):
    session = db.session   
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item.cs_file, data_only=True)
    csSheet = csFile.active
 
    print('--------       Query -|')
    '''
    # Check If a revision for this document already exist
    DRAS_2544-17-DW-0510-04_CY.xlsx
    '''
    try:
        document = item.cs_file.split('_sep_DRAS_')[1].split('_')[0]
        
    except:
        return abort(400, 'Error in your File Name.')
    #doc = session.query(Document).filter(Document.id == item.document_id).first()

    rev = session.query(Drasrevision).filter(Drasrevision.id == item.drasrevision_id).first()
    
        
    rev.stage = item.stage
    
    print(rev.id)
    #print(doc.id)
    session.flush()
    print(rev.id)
    #print(doc.id)

    '''
        HEADER - UPDATE THE COMMENT SHEET
    '''

    
    print(csSheet['C9'].value)
    
    try:
        print('before maybe here')
        #item.revision_id = rev.id
        #item.document_id = item.document_id
        
        
        item.ownerTransmittalReference = csSheet['C9'].value
        
        item.ownerTransmittalDate = date_parse(csSheet['D9'].value)
        print(csSheet['C12'].value)
        item.response_status = csSheet['C12'].value
        
        
        item.contractorTransmittalReference = csSheet['H8'].value
        item.contractorTransmittalDate = date_parse(csSheet['H9'].value)
        

        item.contractorTransmittalMr = csSheet['H10'].value
        item.contractorTransmittalVendor = csSheet['H11'].value
        
        item.documentReferenceDoc = csSheet['K8'].value
        item.documentReferenceRev = csSheet['K9'].value
        item.documentReferenceDesc = csSheet['K10'].value
        item.documentReferenceBy = csSheet['K11'].value
        
        '''
        BODY - CREATE NEW COMMENTS FOR THIS CS
        '''
        
        

        doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
        session.query(Drascomment).filter(Drascomment.drasdocument_id == doc.id).delete()
        print('Document',doc.id, doc.name)
        cs_list = session.query(Drascommentsheet).filter( 
                                    Drascommentsheet.drasdocument_id == doc.id,
                                    Drascommentsheet.current == True).all()
        #item.stage = rev_stage
        for cs in cs_list:  
            cs.current = False
            print('cs', cs.current, cs.id, 'item',item.current, item.id)
        item.current = True
        
        print('Item Current', item.current)
        for row in csSheet.iter_rows(min_row=17,min_col=2):
            
            
            

            if row[0].value is not None:
                #print(row[0].value) 
                comment = Drascomment(

                    drasrevision_id = rev.id,
                    drascommentsheet = item,

                    pos = row[0].value,
                    tag = row[1].value,
                    info = row[2].value,
                    ownerCommentBy = row[3].value,
                    ownerCommentDate = date_parse(csSheet['F15'].value),
                    ownerCommentComment = row[4].value,

                    contractorReplyDate = date_parse(csSheet['H15'].value),
                    contractorReplyStatus = row[5].value,
                    contractorReplyComment = row[6].value,
                    
                    ownerCounterReplyDate = date_parse(csSheet['J15'].value),
                    ownerCounterReplyComment = row[7].value,

                    finalAgreementDate = date_parse(csSheet['L15'].value),
                    finalAgreemntCommentDate = date_parse(row[8].value),
                    finalAgreementComment = row[9].value,

                    commentStatus = row[10].value,
                )
                print('-----         ************       --------')
                print(item.current)
                if item.current == True:
                    
                    comment.drasdocument_id = doc.id
                    
                    print(comment.drasdocument_id, doc.id)
                    

                
                session.add(comment)
        #session.query(Comment).filter(Comment.document_id == doc.id).delete()
        
        print('maybe here')
        session.commit()
        return True
    except:
        abort(400,'Error - Data in Table badly formatted :( - check your file !')


def get_oc(unit, discipline):
    session = db.session
    unit_id = session.query(Unitmodel).filter(Unitmodel.code == unit).first()
    discipline_id = session.query(Disciplinedras).filter(Disciplinedras.name == discipline).first()
    
    if unit_id and discipline_id:
        splitOfWorks = session.query(Splitofworks).filter(
                    Splitofworks.unit_id == unit_id.id,
                    Splitofworks.discipline_id == discipline_id.id).first()
    
        return unit_id.moc_id, splitOfWorks.oc_id
    if unit_id:
        return unit_id.moc_id, unit_id.dedoc_id
    return abort(400,'Unit not Found, check your file name.')

 
def get_data_from_cs(item):
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    
    
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item.cs_file, data_only=True)
    csSheet = csFile.active
 
    print('--------       Query -|')
    '''
    # Check If a revision for this document already exist
    DRAS_2544-17-DW-0510-04_CY.xlsx
    '''
    try:
        document = item.cs_file.split('_sep_DRAS_')[1].split('_')[0]
        full_revision = item.cs_file.split('_sep_DRAS_')[1].split('_')[1].split('.')[0]
        print('Heeeeeeeeeeeere ********************' )
        
        try:
            revision = full_revision[:full_revision.index('S')]
            #rev_stage = full_revision[full_revision.index('S'):]
            rev_stage = 'S'
        except:
            revision = full_revision[:full_revision.index('Y')]
            rev_stage = full_revision[full_revision.index('Y'):]

        oc_unit = document.split('-')[1]
        project = document.split('-')[0] 

    except:
        abort(400, 'Error in file name. Check Your File!')

    doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
    
    if doc is None:
        #fake Discipline
        
        discipline = csSheet['L11'].value

        print('DOC is NONE *-----------           ************')
        moc, dedoc = get_oc(oc_unit, discipline)
        print('MOC - DEDOC ', moc, dedoc)
        doc = Drasdocument(name=document, moc_id=moc, dedoc_id=dedoc)
        session.add(doc)
        print('Document',doc.name)

    # session flush for doc id
    # search the same rev for this document by doc id
    
    rev = session.query(Drasrevision).filter(Drasrevision.name == revision, Drasrevision.drasdocument_id == doc.id).first() 
    print('searching for revision, document:', revision, document)
    print('found', rev)
    if rev is None:
        print(rev)
        print('    ----------     Rev is None: ', revision, rev_stage, document)
        rev = Drasrevision(name=revision, drasdocument=doc)
        session.add(rev)
        
    rev.stage = rev_stage
    
     
    session.flush()

    '''
        HEADER - UPDATE THE COMMENT SHEET
    '''

    item.drasrevision_id = rev.id
    item.drasdocument_id = doc.id

    item.ownerTransmittalReference = csSheet['C9'].value
    item.ownerTransmittalDate = date_parse(csSheet['D9'].value)
    item.response_status = csSheet['C12'].value

    item.contractorTransmittalReference = csSheet['H8'].value
    item.contractorTransmittalDate = date_parse(csSheet['H9'].value)
    item.contractorTransmittalMr = csSheet['H10'].value
    item.contractorTransmittalVendor = csSheet['H11'].value

    item.documentReferenceDoc = csSheet['L8'].value
    item.documentReferenceRev = csSheet['L9'].value
    item.documentReferenceDesc = csSheet['L10'].value
    
    # Discipline
    item.documentReferenceBy = csSheet['L11'].value

    #item.documentReferenceBy = fdiscipline
    indoor = ['Y','Y2']
    outdoor = ['S','Y1','Y3']

    if rev_stage in indoor: 
        item.expectedDate = item.actualDate + timedelta(days=7)
    if rev_stage in outdoor:
        item.expectedDate = item.actualDate + timedelta(days=14)
    
    '''
    BODY - CREATE NEW COMMENTS FOR THIS CS
    '''
    
    if item.current:
        session.query(Drascomment).filter(Drascomment.drasdocument_id == doc.id).delete()
        
        commentSheets = session.query(Drascommentsheet).filter(Drascommentsheet.drasdocument_id == doc.id).all()
        item.stage = rev_stage

        for cs in commentSheets:
            cs.current = False
    

    try:
        for row in csSheet.iter_rows(min_row=17,min_col=2):
            print('CommentStatus', row[0].value,row[9].value,row[10].value,row[11].value, type(row[11].value))
            
            if row[0].value is not None and row[1].value is not None:
                print(row[0].value)
                #  
                comment = Drascomment(
                    drasrevision_id = rev.id,
                    drascommentsheet = item,
                    tagdiscipline= session.query(Tagdiscipline).filter(
                                                Tagdiscipline.start <= int(row[1].value), 
                                                Tagdiscipline.finish >= int(row[1].value)).first(), 

                    pos = row[0].value,
                    tag = row[1].value,
                    info = row[2].value,
                    ownerCommentBy = row[3].value,
                    ownerCommentDate = date_parse(csSheet['F15'].value),
                    ownerCommentComment = row[4].value,

                    contractorReplyDate = date_parse(csSheet['H15'].value),
                    contractorReplyStatus = row[5].value,
                    contractorReplyComment = row[6].value,
                    
                    ownerCounterReplyDate = date_parse(csSheet['K15'].value),
                    ownerCounterReplyComment = row[8].value,

                    finalAgreementDate = date_parse(csSheet['M15'].value),
                    finalAgreemntCommentDate = date_parse(row[10].value),
                    finalAgreementComment = row[11].value,

                    commentStatus = str(row[12].value),
                )

                if item.current:
                    print('BLOCKED HERE ------------------ //////////////////////')
  
                    
                    comment.drasdocument_id = doc.id
                    

                #print('Contractor Status:',len(comment.contractorReplyStatus),comment.contractorReplyStatus)
                session.add(comment)
        #session.query(Comment).filter(Comment.document_id == doc.id).delete()

        session.commit()
        return doc.id

    
    except:
        abort(400,'Error - Data in Table badly formatted :( - check your file !')
     


def get_fake_data_from_cs2(item):
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    
    
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+'fakeDras/DRAS_2544-13-MOM-4561-09_A0Y.xlsx', data_only=True)
    csSheet = csFile.active 
 
    #print('--------       Query -|')
    '''
    # Check If a revision for this document already exist
    DRAS_2544-17-DW-0510-04_CY.xlsx
    '''
    try: 
        document = item.cs_file.split('_sep_DRAS_')[1].split('_')[0]
        full_revision = item.cs_file.split('_sep_DRAS_')[1].split('_')[1].split('.')[0]
        revision = full_revision[:full_revision.index('Y')]
        rev_stage = full_revision[full_revision.index('Y'):]

        oc_unit = document.split('-')[1]
        project = document.split('-')[0] 

    except:
        abort(400, 'Error in file name. Check Your File!')

    doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
    
    if doc is None:
        
        discipline = csSheet['K11'].value
        #print('DOC is NONE *-----------           ************')
        moc, dedoc = get_oc(oc_unit, discipline)
        print('MOC - DEDOC ', moc, dedoc)
        doc = Drasdocument(name=document, 
                        moc_id=moc, 
                        dedoc_id=dedoc,
                        created_by_fk='1',
                        changed_by_fk='1')
        session.add(doc)
        session.flush()
        #print('Document',doc.name)

    # session flush for doc id
    # search the same rev for this document by doc id

    rev = session.query(Drasrevision).filter(Drasrevision.name == revision, Drasrevision.drasdocument_id == doc.id).first() 
    print('searching for revision, document:', revision,rev_stage, document)
    #print('found', rev)
    #rev.stage = rev_stage
    #rev.changed_by_fk = '1'
    if rev:
        rev.stage = rev_stage
        rev.changed_by_fk = '1'
    else:
        #print(rev)
        #print('    ----------     Rev is None: ', revision, rev_stage, document)
        rev = Drasrevision(name=revision, drasdocument=doc,
                created_by_fk='1',
                changed_by_fk='1',
                stage=rev_stage)
        
        session.add(rev)
        session.flush()
    
    #rev.stage = rev_stage
    #rev.changed_by_fk = '1'
    
    
    
    
    

    '''
        HEADER - UPDATE THE COMMENT SHEET
    '''

    item.drasrevision_id = rev.id
    item.drasdocument_id = doc.id

    item.ownerTransmittalReference = csSheet['C9'].value
    item.ownerTransmittalDate = date_parse(csSheet['D9'].value)
    item.response_status = csSheet['C12'].value

    item.contractorTransmittalReference = csSheet['H8'].value
    item.contractorTransmittalDate = date_parse(csSheet['H9'].value)
    item.contractorTransmittalMr = csSheet['H10'].value
    item.contractorTransmittalVendor = csSheet['H11'].value

    item.documentReferenceDoc = csSheet['K8'].value
    item.documentReferenceRev = csSheet['K9'].value
    item.documentReferenceDesc = csSheet['K10'].value
    item.documentReferenceBy = csSheet['K11'].value
    
    
    '''
    BODY - CREATE NEW COMMENTS FOR THIS CS
    '''
    
    if item.current:
        session.query(Drascomment).filter(Drascomment.drasdocument_id == doc.id).delete()
        
        commentSheets = session.query(Drascommentsheet).filter(Drascommentsheet.drasdocument_id == doc.id).all()
        item.stage = rev_stage

        for cs in commentSheets:
            cs.current = False
            cs.changed_by_fk = '1'
    
    # random comment status
    #
    #   
    
    
    def random_status():
        status = ['Open', 'Closed']
        return status[randint(0,1)]
        
    try:
        for row in csSheet.iter_rows(min_row=17,min_col=2):
            #print('CommentStatus', row[0].value,row[9].value,row[10].value,row[11].value, type(row[11].value))
            if row[0].value is not None:
                #print(row[0].value) 
                comment = Drascomment(
                    drasrevision_id = rev.id,
                    drascommentsheet = item,

                    pos = row[0].value,
                    tag = row[1].value,
                    info = row[2].value,
                    ownerCommentBy = row[3].value,
                    ownerCommentDate = date_parse(csSheet['F15'].value),
                    ownerCommentComment = row[4].value,

                    contractorReplyDate = date_parse(csSheet['H15'].value),
                    contractorReplyStatus = row[5].value,
                    contractorReplyComment = row[6].value,
                    
                    ownerCounterReplyDate = date_parse(csSheet['J15'].value),
                    ownerCounterReplyComment = row[7].value,

                    finalAgreementDate = date_parse(csSheet['L15'].value),
                    finalAgreemntCommentDate = date_parse(row[9].value),
                    finalAgreementComment = row[10].value,

                    #commentStatus = str(row[11].value),
                    commentStatus = random_status(),

                    created_by_fk='1',
                    changed_by_fk='1'
                )
                if item.current:  
                    
                    comment.drasdocument_id = doc.id
                    

                #print('Contractor Status:',len(comment.contractorReplyStatus),comment.contractorReplyStatus)
                session.add(comment)
        #session.query(Comment).filter(Comment.document_id == doc.id).delete()

        session.commit()
        return doc.id

    
    except:
        abort(400,'Error - Data in Table badly formatted :( - check your file !')
     

def get_data_from_cs3(item):
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    
    
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item.cs_file, data_only=True)
    csSheet = csFile.active
 
    print('--------       Query -|')
    '''
    # Check If a revision for this document already exist
    DRAS_2544-17-DW-0510-04_CY.xlsx
    '''
    try:
        document = item.cs_file.split('_sep_DRAS_')[1].rsplit('_',1)[0]
        full_revision = item.cs_file.split('_sep_DRAS_')[1].rsplit('_',1)[1].split('.')[0]
        print('Heeeeeeeeeeeere ********************',document,full_revision )
        
        try:
            print('First Try on filename')
            revision = full_revision[:full_revision.index('S')]
            rev_stage = full_revision[full_revision.index('S'):]
            print('First Try on filename', revision, rev_stage)
        except:
            print('First Exception on filename')
            revision = full_revision[:full_revision.index('Y')]
            rev_stage = full_revision[full_revision.index('Y'):]
            print('First Exception on filename', revision, rev_stage)

        oc_unit = document.split('-')[1]
        project = document.split('-')[0] 

    except:
        abort(400, 'Check: 001 | Il nome del file è errato.')

    doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
    
    if doc is None:
        #fake Discipline
        
        discipline = csSheet['L11'].value

        print('DOC is NONE *-----------           ************')
        moc, dedoc = get_oc(oc_unit, discipline)
        print('MOC - DEDOC ', moc, dedoc)
        doc = Drasdocument(name=document, moc_id=moc, dedoc_id=dedoc)
        session.add(doc)
        print('Document',doc.name)

    # session flush for doc id
    # search the same rev for this document by doc id
    
    rev = session.query(Drasrevision).filter(Drasrevision.name == revision, Drasrevision.drasdocument_id == doc.id).first() 
    print('searching for revision, document:', revision, document)
    print('found', rev)
    if rev is None:
        print(rev)
        print('    ----------     Rev is None: ', revision, rev_stage, document)
        rev = Drasrevision(name=revision, drasdocument=doc)
        rev.stage = rev_stage
        session.add(rev)
        
    rev.stage = rev_stage
    
   
    #Check If this DRAS STAGE already exist
    ds = session.query(Drascommentsheet).filter(
        Drascommentsheet.drasdocument_id == doc.id,
        Drascommentsheet.drasrevision_id == rev.id,
        Drascommentsheet.stage == rev_stage
    ).first()
    if ds:
        print(doc.id, rev.id, rev_stage)
        flash('This DRAS Stage already Exist', category='info')

        abort(409)

    '''
    if rev_stage == 'S' and not item.actionrequired_id and not item.issuetype_id:
        print(item.actionrequired_id, item.issuetype)
        flash('DRAS in Stage "S", please select Action e Issue Type', category='warning')

        abort(409)
    '''
    session.flush()

    '''
        HEADER - UPDATE THE COMMENT SHEET
    '''

    item.drasrevision_id = rev.id
    item.drasdocument_id = doc.id

    item.ownerTransmittalReference = csSheet['C9'].value
    item.ownerTransmittalDate = date_parse(csSheet['D9'].value)
    item.response_status = csSheet['C12'].value

    item.contractorTransmittalReference = csSheet['H8'].value
    item.contractorTransmittalDate = date_parse(csSheet['H9'].value)
    item.contractorTransmittalMr = csSheet['H10'].value
    item.contractorTransmittalVendor = csSheet['H11'].value

    item.documentReferenceDoc = csSheet['L8'].value
    item.documentReferenceRev = csSheet['L9'].value
    item.documentReferenceDesc = csSheet['L10'].value
    
    # Discipline
    item.documentReferenceBy = csSheet['L11'].value

    #item.documentReferenceBy = fdiscipline
    indoor = ['S','Y','Y2']
    outdoor = ['Y1','Y3']

    if rev_stage in indoor: 
        item.expectedDate = item.actualDate + timedelta(days=14)
    if rev_stage in outdoor:
        item.expectedDate = item.actualDate + timedelta(days=7)
    
    '''
    BODY - CREATE NEW COMMENTS FOR THIS CS
    '''
    
    if item.current:
        session.query(Drascomment).filter(Drascomment.drasdocument_id == doc.id).delete()
        
        commentSheets = session.query(Drascommentsheet).filter(Drascommentsheet.drasdocument_id == doc.id).all()
        item.stage = rev_stage

        for cs in commentSheets:
            cs.current = False
    
    item.stage = rev_stage
    session.commit()
    try:
        none_count = 0
        for row in csSheet.iter_rows(min_row=17,min_col=2):
            print('CommentStatus', row[0].value,row[9].value,row[10].value,row[11].value, type(row[11].value))
            
            if row[0].value is not None and row[1].value is not None:
                print(row[0].value)
                #  
                comment = Drascomment(
                    drasrevision_id = rev.id,
                    drascommentsheet = item,
                    tagdiscipline= session.query(Tagdiscipline).filter(
                                                Tagdiscipline.start <= int(row[1].value), 
                                                Tagdiscipline.finish >= int(row[1].value)).first(), 

                    pos = row[0].value,
                    tag = row[1].value,
                    info = row[2].value,
                    ownerCommentBy = row[3].value,
                    ownerCommentDate = date_parse(csSheet['F15'].value),
                    ownerCommentComment = row[4].value,

                    contractorReplyDate = date_parse(csSheet['H15'].value),
                    contractorReplyStatus = row[5].value,
                    contractorReplyComment = row[6].value,
                    
                    ownerCounterReplyDate = date_parse(csSheet['K15'].value),
                    ownerCounterReplyComment = row[8].value,

                    finalAgreementDate = date_parse(csSheet['M15'].value),
                    finalAgreemntCommentDate = date_parse(row[10].value),
                    finalAgreementComment = row[11].value,

                    commentStatus = str(row[12].value),
                )

                if item.current:
                    print('BLOCKED HERE ------------------ //////////////////////')
  
                    
                    comment.drasdocument_id = doc.id
                    

                #print('Contractor Status:',len(comment.contractorReplyStatus),comment.contractorReplyStatus)
            
                session.add(comment)
                
            else:
                # Check on inifinite excel issue
                none_count += 1
                if none_count > 25:
                    flash('Excel BAD FORMAT: Infinite Comments', category='info')
                    raise Exception('Excel BAD FORMAT: Infinite Comments')
            
            
            
        #session.query(Comment).filter(Comment.document_id == doc.id).delete() 
    except:
        session.rollback()
        
        flash('COMMENTS ERROR 003 | Non è stato possibile caricare i commenti per questo DRAS', category='warning')
        item.note = 'COMMENTS ERROR 003: Badly Formatted. Please find the attached original DRAS in order to review you comments.'

    
    
    session.commit()
    return doc.id



'''
def find_rev():
    session = db.session
    document = '2544-17-DW-0510-04'
    revision = 'CYF'

    rev = session.query(Revision).filter(Revision.name == revision, Document.name == document).first()
    if rev is None:
        print('     ----- Rev is NONE')
    print('    ------  RR rev: ',rev.id, rev.name, rev.document)
    
    return
'''
#find_rev()

def set_current_last_actual_date():
    session = db.session
    docs = session.query(Drasdocument).all()
    bad_file = []
    
    for doc in docs:
        try:
            cs_list = session.query(Drascommentsheet).filter(
                Drascommentsheet.drasdocument_id == doc.id
            ).all()
            print(cs_list)
            for c in cs_list:
                print(c.id)
                c.changed_by_fk = '1'
                c.created_by_fk = '1'
                c.current = False
            session.commit()    
                
            
            cs = session.query(Drascommentsheet).filter(
                Drascommentsheet.drasdocument_id == doc.id
            ).order_by(Drascommentsheet.actualDate.desc()).first()
            print(doc,cs.id)
            cs.current = True
            cs.changed_by_fk = '1'
            cs.created_by_fk = '1'
        except:
            bad_file.append(cs)      
            
    session.commit()

#set_current_last_actual_date() 
    # 
    #  
def set_expected_date():
    session = db.session
    cs_list = session.query(Drascommentsheet).all()
    indoor = ['Y','Y2']
    outdoor = ['Y1', 'Y3','S']
    for cs in cs_list:
        try:

            print(cs.id)
            if cs.actualDate:
                if cs.stage in indoor:
                    cs.expectedDate = cs.actualDate + timedelta(days=7)
                elif cs.stage in outdoor:
                    cs.expectedDate = cs.actualDate + timedelta(days=15)
            cs.changed_by_fk = '1'
            cs.created_by_fk = '1'
        except:
            print('SOMETHING WRONG')
        
           
    session.commit()

#set_expected_date() 
        



import os
from random import choice, randrange
from time import sleep
def check_labels2(item):
    # 
    # Check File Labels
    # 
    #  
    #item = item.cs_file
    csFile = openpyxl.load_workbook(item)
    csSheet = csFile.active
    
    # Duplicate key (label in excel) for check purpose
    header_labels_dict = {
        'Reference' : csSheet['C8'].value,
        'Date' : csSheet['D8'].value,

        'Reference' : csSheet['G8'].value,
        'Date' : csSheet['G9'].value,
        'Material requisition': csSheet['G10'].value,
        'Vendor Name' : csSheet['G11'].value,

        'Rev.' : csSheet['K9'].value,
        'Description': csSheet['K10'].value,
        'Issued by (Contractor Discipline)': csSheet['K11'].value
    }

    for key, value in header_labels_dict.items():
        #print(key, value)
        if key != value:
            #flash(('Header Label ' + key + ' Not Found, Check your DRAS Template!'), category='warning')
            abort(400,('Header Label ' + key + ' Not Found, Check your DRAS Template!'))
    
    table_label_dict = {
        'Pos.' : csSheet['B14'].value,
        
        'Status' : csSheet['G16'].value,

        'Date' : csSheet['L15'].value,

    }

    for key, value in table_label_dict.items():
        #print(key, value)
        if key != value:
            flash(('Table Label ' + key + ' Not Found, Check your DRAS Template!'), category='warning')
            print(('Table Label ' + key + ' Not Found, Check your DRAS Template!'))
            return False
    
    print('-------------------- CHECK FUNCTION TRUE')
    return True

def cs_data_report(dras_file):
        
    print(dras_file)
    cs_file = dras_file.split('/')[-1]
    print(cs_file)
    try:
        print('BEFORE ********************' )
        document = cs_file.split('DRAS_')[1].split('_')[0]
        

        full_revision = cs_file.split('DRAS_')[1].split('_')[1].split('.')[0]
        
        
        try:
            revision = full_revision[:full_revision.index('S')]
            rev_stage = full_revision[full_revision.index('S'):]
        except:
            
            revision = full_revision[:full_revision.index('Y')]
            rev_stage = full_revision[full_revision.index('Y'):]
        
        oc_unit = document.split('-')[1]
        project = document.split('-')[0]
        print(document,revision,rev_stage,oc_unit,project)

    except:
        abort(400, 'Error in file name. Check Your File!')


    
    issue_type = 'it'
    action = 'ac'
    not_item = 'ni'
    actual_date = 'ad'
    
    return issue_type, action, not_item, actual_date

def batch_upload():
    session = db.session
    users = [1,2,3]
    batch_folder = 'init/dras_s'
    #batch_folder = 'settimo_batch'
    path = UPLOAD_FOLDER + batch_folder

    files = []
    # r=root, d=directories, f = files
    for r, d, f in os.walk(path):
        for file in f:
            if '.xlsx' in file:
                files.append(os.path.join(r, file))

    bad_file = []
    for f in files:
        try:
            print(f)
            
            '''
            issue_type, action, not_item, actual_date = cs_data_report(f)
        
            print(issue_type, action, not_item, actual_date)
            print(' BEFORE LABELS ********************' )
            check_labels2(f)
            print(' LABELS ********************' )
            '''
            csFile = openpyxl.load_workbook(f)
            csSheet = csFile.active

            notification_item = csSheet['H8'].value
            actual_date = csSheet['H9'].value

            full_file = f.split('/')[-1]
            document = full_file.split('_')[1]
            revision = full_file.split('_')[2].split('.')[0]

            print(document, revision)
            oc_unit = document.split('-')[1]
            project = document.split('-')[0]

            rev_stage = 'S'
            doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()

            if doc is None:
                #fake Discipline
                
                discipline = csSheet['L11'].value

                print('DOC is NONE *-----------           ************')
                moc, dedoc = get_oc(oc_unit, discipline)
                print('MOC - DEDOC ', moc, dedoc)
                doc = Drasdocument(name=document, moc_id=moc, dedoc_id=dedoc)
                doc.created_by_fk = '1'
                doc.changed_by_fk = '1'
                print('DOC           ************')
                session.add(doc)
                print('DOC  ADD         ************')
                session.flush()
                print('DOC  FLUSH         ************')
                print('Document',doc.name)
                print(revision, doc.id)
            # session flush for doc id
            session.flush()
            print('AFTER FLUSH')
            print(revision, doc.id)
            # search the same rev for this document by doc id
            
            print('BLOCKED HERE ------------------ //////////////////////')
            rev = session.query(Drasrevision).filter(Drasrevision.name == revision, Drasrevision.drasdocument_id == doc.id).first() 
            print('searching for revision, document:', revision, document)
            print('found', rev)
            if rev is None:
                print(rev)
                print('    ----------     Rev is None: ', revision, rev_stage, document)
                rev = Drasrevision(name=revision, drasdocument=doc, changed_by_fk='1',created_by_fk = '1')
                
                session.add(rev)
                #session.flush()
            
            rev.changed_by_fk = '1'
            rev.stage = rev_stage
            #rev.created_by_fk = '1'
            
            
            
            session.flush()
            

            '''
                HEADER - UPDATE THE COMMENT SHEET
            '''

                

            '''
            except:
                print('FAILED XXXXXXXXXXXXXXXXXXXXXXXXXX ',f)
                bad_file.append(f)
            '''

            #issue_type, action, not_item, actual_date = cs_data_report(f)
            user = choice(users)
            if isinstance(actual_date,str):
                try:
                    actual_date = datetime.strptime(actual_date,'%d/%m/%Y')
                except:
                    print('FAILED######################### ',f)
                    bad_file.append(f)
            
            drascs = Drascommentsheet(
                created_by_fk = user,
                changed_by_fk = user,
                cs_file = f,
                #issuetype = issue_type,
                #actionrequired = action,
                actualDate = actual_date,
                #actualDate = actual_date,
                notificationItem = notification_item,
                stage = 'S'
            )
            drascs.drasrevision_id = rev.id
            drascs.drasdocument_id = doc.id
            drascs.created_by_fk = '1'
            drascs.changed_by_fk = '1'
            session.add(drascs)
        except:
            bad_file.append(f)
    session.commit()
    print('***********BAD FILES LIST',len(bad_file))
    print(bad_file)

#batch_upload()  

def update_discipline():
    session = db.session
    users = [1,2,3]
    batch_folder = 'init/dras_s'
    #batch_folder = 'settimo_batch'
    path = UPLOAD_FOLDER + batch_folder

    files = []
    # r=root, d=directories, f = files
    for r, d, f in os.walk(path):
        for file in f:
            if '.xlsx' in file:
                files.append(os.path.join(r, file))

    bad_file = []
    for f in files:
        
        try:
            print(f)
            
            '''
            issue_type, action, not_item, actual_date = cs_data_report(f)
        
            print(issue_type, action, not_item, actual_date)
            print(' BEFORE LABELS ********************' )
            check_labels2(f)
            print(' LABELS ********************' )
            '''
            csFile = openpyxl.load_workbook(f)
            csSheet = csFile.active

            discipline = csSheet['L11'].value
            description = csSheet['L10'].value
            cs_rev = csSheet['L9'].value
            cs_doc = csSheet['L8'].value
            

            full_file = f.split('/')[-1]
            document = full_file.split('_')[1]
            revision = full_file.split('_')[2].split('.')[0]
            stage = 'S'

            print(document, revision)

            doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
            print('DOC ID', doc.id)
            rev = session.query(Drasrevision).filter(Drasrevision.drasdocument_id == doc.id,
                                Drasrevision.name == revision).first()
            
            print('FOUND +++++ =',doc.name, rev.name)
            cs = session.query(Drascommentsheet).filter(
                Drascommentsheet.drasdocument_id == doc.id,
                Drascommentsheet.drasrevision_id == rev.id,
                Drascommentsheet.stage == stage
            ).first()
            cs.documentReferenceBy = discipline[:49]
            cs.documentReferenceDesc = description
            cs.documentReferenceRev = cs_rev
            cs.documentReferenceDoc = cs_doc

            cs.changed_by_fk = '1'
            print(cs.id,cs.documentReferenceBy,cs.documentReferenceDoc,cs.documentReferenceRev)
            session.commit()
        except:
            print('Something Wrong')
        

#update_discipline() 

def set_dsc():
    ''' Include all Janus Milestone '''
    session = db.session
    # Open the MDI Template
    set_dsc = open('xlsx/set_dsc.xlsx', mode='rb')
    wmb = openpyxl.load_workbook(set_dsc, data_only=True)
    ws = wmb.active

    for row in ws.iter_rows(min_row=1):
        print(row[0].value, row[1].value)
        try:
            cs = session.query(Drascommentsheet).filter(Drascommentsheet.id == row[0].value).first()
            cs.documentReferenceBy = row[1].value
            cs.changed_by_fk = '1'
        except:
            print('no document found')

    session.commit()

#set_dsc()
def set_doc_dsc():
    ''' Set the document discipline for all DRASs '''
    session = db.session
    
    set_dsc = open('xlsx/set_doc_dsc.xlsx', mode='rb')
    wmb = openpyxl.load_workbook(set_dsc, data_only=True)
    ws = wmb.active

    for row in ws.iter_rows(min_row=1):
        print(row[0].value, row[1].value)
        try:
            doc_code = row[1].value
            discipline = row[0].value
            doc = session.query(Drasdocument).filter(Drasdocument.name == doc_code).first()
            cs_list = session.query(Drascommentsheet).filter(Drascommentsheet.drasdocument_id == doc.id).all()
            for cs in cs_list:
                print(cs.changed_by_fk, cs.documentReferenceBy)
                cs.documentReferenceBy = discipline
                cs.changed_by_fk = '1'
                #doc.changed_by_fk = '1'
                session.commit()
                #session.add(cs)
                print(cs.changed_by_fk, cs.documentReferenceBy)
        except:
            print('no document found')

    #session.commit()
#set_doc_dsc() 
def set_process():
    session = db.session
   
    cs_list = session.query(Drascommentsheet).filter(Drascommentsheet.documentReferenceBy == 'Process').all() 
    print('               ********************       ***** ')
    print('Process list,', len(cs_list))
    for cs in cs_list:
        print('something') 
        cs.documentReferenceBy = 'PROCESS MANAGER/COORDINATOR'
        cs.changed_by_fk = '1'
    session.commit()

#set_process() 
  

def set_current_and_last(item):
    db_session = db.session
    bad_file = []

    '''SET ALL CS FOR THIS DOCUMENT TO FALSE CURRENT AND LAST STAGE '''
    all_cs = db_session.query(Drascommentsheet).filter(
        Drascommentsheet.drasdocument_id == item.drasdocument_id
    ).all
    for cs in all_cs:
        cs.current = False
        cs.last_stage = False

    '''SET TRUE LAST STAGE FOR REVISION'''
    revisions = db_session.query(Drasrevision).filter(
        Drasrevision.drasdocument_id == item.drasdocument_id
    ).all()
    for rev in revisions:
        try:
            cs_last_stage = db_session.query(Drascommentsheet).filter(
                Drascommentsheet.drasdocument_id == item.drasdocument_id,
                Drascommentsheet.drasrevision_id == rev.id
            ).order_by(Drascommentsheet.actualDate.desc()).first()
            #print(cs_list)
            cs_last_stage.last_stage = True
            db_session.commit()     
        except:
            bad_file.append(item) 
        try:     
            '''SET TRUE CURRENT FOR DOCUMENT'''
            cs = db_session.query(Drascommentsheet).filter(
                    Drascommentsheet.drasdocument_id == item.drasdocument_id
                ).order_by(Drascommentsheet.actualDate.desc()).first()
            print(item.drasdocument_id,cs.id)
            cs.current = True
            #cs.changed_by_fk = '1'
            #cs.created_by_fk = '1'
        except:
            bad_file.append(item) 
    db_session.commit()

def get_data_from_cs4(item):
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    
    
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item.cs_file, data_only=True)
    csSheet = csFile.active
 
    print('--------       Query -|')
    '''
    # Check If a revision for this document already exist
    DRAS_2544-17-DW-0510-04_CY.xlsx
    '''
    try:
        document = item.cs_file.split('_sep_DRAS_')[1].rsplit('_',1)[0]
        full_revision = item.cs_file.split('_sep_DRAS_')[1].rsplit('_',1)[1].split('.')[0]
        print('Heeeeeeeeeeeere ********************',document,full_revision )
        
        try:
            print('First Try on filename')
            revision = full_revision[:full_revision.index('S')]
            rev_stage = full_revision[full_revision.index('S'):]
            print('First Try on filename', revision, rev_stage)
        except:
            print('First Exception on filename')
            revision = full_revision[:full_revision.index('Y')]
            rev_stage = full_revision[full_revision.index('Y'):]
            print('First Exception on filename', revision, rev_stage)

        oc_unit = document.split('-')[1]
        project = document.split('-')[0] 

    except:
        abort(400, 'Check: 001 | Il nome del file è errato.')

    doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
    
    if doc is None:
        #fake Discipline
        if csSheet['L14'].value == 'Final Agreement':
            discipline = csSheet['L11'].value
        else:
            discipline = csSheet['K11'].value

        print('DOC is NONE *-----------           ************')
        moc, dedoc = get_oc(oc_unit, discipline)
        print('MOC - DEDOC ', moc, dedoc)
        doc = Drasdocument(name=document, moc_id=moc, dedoc_id=dedoc)
        session.add(doc)
        print('Document',doc.name)

    # session flush for doc id
    # search the same rev for this document by doc id
    
    rev = session.query(Drasrevision).filter(Drasrevision.name == revision, Drasrevision.drasdocument_id == doc.id).first() 
    print('searching for revision, document:', revision, document)
    print('found', rev)
    if rev is None:
        print(rev)
        print('    ----------     Rev is None: ', revision, rev_stage, document)
        rev = Drasrevision(name=revision, drasdocument=doc)
        rev.stage = rev_stage
        session.add(rev)
        
    rev.stage = rev_stage
    
   
    #Check If this DRAS STAGE already exist
    ds = session.query(Drascommentsheet).filter(
        Drascommentsheet.drasdocument_id == doc.id,
        Drascommentsheet.drasrevision_id == rev.id,
        Drascommentsheet.stage == rev_stage
    ).first()
    if ds:
        print(doc.id, rev.id, rev_stage)
        flash('This DRAS Stage already Exist', category='info')

        abort(409)

    '''
    if rev_stage == 'S' and not item.actionrequired_id and not item.issuetype_id:
        print(item.actionrequired_id, item.issuetype)
        flash('DRAS in Stage "S", please select Action e Issue Type', category='warning')

        abort(409)
    '''
    session.flush()

    '''
        HEADER - UPDATE THE COMMENT SHEET
    '''
    item.drasrevision_id = rev.id
    item.drasdocument_id = doc.id

    if csSheet['L14'].value == 'Final Agreement':
        item.ownerTransmittalReference = csSheet['C9'].value
        item.ownerTransmittalDate = date_parse(csSheet['D9'].value)
        item.response_status = csSheet['C12'].value

        item.contractorTransmittalReference = csSheet['H8'].value
        item.contractorTransmittalDate = date_parse(csSheet['H9'].value)
        item.contractorTransmittalMr = csSheet['H10'].value
        item.contractorTransmittalVendor = csSheet['H11'].value

        item.documentReferenceDoc = csSheet['L8'].value
        item.documentReferenceRev = csSheet['L9'].value
        item.documentReferenceDesc = csSheet['L10'].value
        
        # Discipline
        item.documentReferenceBy = csSheet['L11'].value
    else:
        item.ownerTransmittalReference = csSheet['C9'].value
        item.ownerTransmittalDate = date_parse(csSheet['D9'].value)
        item.response_status = csSheet['C12'].value

        item.contractorTransmittalReference = csSheet['H8'].value
        item.contractorTransmittalDate = date_parse(csSheet['H9'].value)
        item.contractorTransmittalMr = csSheet['H10'].value
        item.contractorTransmittalVendor = csSheet['H11'].value

        item.documentReferenceDoc = csSheet['K8'].value
        item.documentReferenceRev = csSheet['K9'].value
        item.documentReferenceDesc = csSheet['K10'].value
        
        # Discipline
        item.documentReferenceBy = csSheet['K11'].value

    #item.documentReferenceBy = fdiscipline
    indoor = ['S','Y','Y2']
    outdoor = ['Y1','Y3']

    if rev_stage in indoor: 
        item.expectedDate = item.actualDate + timedelta(days=14)
    if rev_stage in outdoor:
        item.expectedDate = item.actualDate + timedelta(days=7)
    
    '''
    BODY - CREATE NEW COMMENTS FOR THIS CS
    '''
    '''
    if item.current:
        session.query(Drascomment).filter(Drascomment.drasdocument_id == doc.id).delete()
        
        commentSheets = session.query(Drascommentsheet).filter(Drascommentsheet.drasdocument_id == doc.id).all()
        item.stage = rev_stage

        for cs in commentSheets:
            cs.current = False
    '''
    item.stage = rev_stage
    item.note = "Piero003"
    session.commit()
    try:
        none_count = 0
        for row in csSheet.iter_rows(min_row=17,min_col=2):
            print('CommentStatus', row[0].value,row[9].value,row[10].value,row[11].value, type(row[11].value))
            
            if row[0].value is not None and row[1].value is not None:
                print(row[0].value)
                # 
                if csSheet['L14'].value == 'Final Agreement': 
                    comment = Drascomment(
                        drasdocument_id = doc.id,
                        drasrevision_id = rev.id,
                        drascommentsheet = item,
                        tagdiscipline= session.query(Tagdiscipline).filter(
                                                    Tagdiscipline.start <= int(row[1].value), 
                                                    Tagdiscipline.finish >= int(row[1].value)).first(), 

                        pos = row[0].value,
                        tag = row[1].value,
                        info = row[2].value,
                        ownerCommentBy = row[3].value,
                        ownerCommentDate = date_parse(csSheet['F15'].value),
                        ownerCommentComment = text_decode(row[4].value),

                        contractorReplyDate = date_parse(csSheet['H15'].value),
                        contractorReplyStatus = row[5].value,
                        contractorReplyComment = text_decode(row[6].value),
                        
                        ownerCounterReplyDate = date_parse(csSheet['K15'].value),
                        ownerCounterReplyComment = text_decode(row[8].value),

                        finalAgreementDate = date_parse(csSheet['M15'].value),
                        finalAgreemntCommentDate = date_parse(row[10].value),
                        finalAgreementComment =text_decode(row[11].value) ,

                        commentStatus = str(row[12].value),
                        
                    )      
                    #print('Contractor Status:',len(comment.contractorReplyStatus),comment.contractorReplyStatus)
                
                    #session.add(comment)
                else:
                    comment = Drascomment(
                        drasdocument_id = doc.id,
                        drasrevision_id = rev.id,
                        drascommentsheet = item,
                        tagdiscipline= session.query(Tagdiscipline).filter(
                                                    Tagdiscipline.start <= int(row[1].value), 
                                                    Tagdiscipline.finish >= int(row[1].value)).first(), 

                        pos = row[0].value,
                        tag = row[1].value,
                        info = row[2].value,
                        ownerCommentBy = row[3].value,
                        ownerCommentDate = date_parse(csSheet['F15'].value),
                        ownerCommentComment = text_decode(row[4].value),

                        contractorReplyDate = date_parse(csSheet['H15'].value),
                        contractorReplyStatus = row[5].value,
                        contractorReplyComment = text_decode(row[6].value),
                        
                        ownerCounterReplyDate = date_parse(csSheet['J15'].value),
                        ownerCounterReplyComment = text_decode(row[7].value),

                        finalAgreementDate = date_parse(csSheet['L15'].value),
                        finalAgreemntCommentDate = date_parse(row[9].value),
                        finalAgreementComment = text_decode(row[10].value),

                        commentStatus = str(row[11].value),
                        
                    )      
                #print('Contractor Status:',len(comment.contractorReplyStatus),comment.contractorReplyStatus)
                ## SET MAX LENGHT
                if comment.commentStatus:
                    comment.commentStatus = comment.commentStatus[:20]
                if comment.ownerCommentBy:
                    comment.ownerCommentBy = comment.ownerCommentBy[:50]
                
                if comment.pos:
                    comment.pos = str(comment.pos)[:5]

                if comment.contractorReplyStatus:
                    comment.contractorReplyStatus = comment.contractorReplyStatus[:100]
                
                session.add(comment)
                
                
            else:
                # Check on inifinite excel issue
                none_count += 1
                if none_count > 35:
                    flash('Excel BAD FORMAT: Infinite Comments', category='info')
                    raise Exception('Excel BAD FORMAT: Infinite Comments')
            
            
        item.comments_uploaded = True   
        #session.query(Comment).filter(Comment.document_id == doc.id).delete() 
    except Exception as e:
        session.rollback()
        
        flash('COMMENTS ERROR 003 | Non è stato possibile caricare i commenti per questo DRAS', category='warning')
        #item.note = 'COMMENTS ERROR 003: Badly Formatted. Please find the attached original DRAS in order to review you comments.'
        
        #flash('COMMENTS ERROR 003 | Non è stato possibile caricare i commenti per questo DRAS', category='warning')
        '''
        error_encode = str(e).encode("ascii", "ignore")
        string_decode = error_encode.decode()
        item.note = str(string_decode)
        '''
        item.note = text_decode(str(e))
        
        item.comments_uploaded = False
    
    
    session.commit()
    return doc.id

def update_current_and_last(item):
    db_session = db.session
    
    cs_list = db_session.query(Drascommentsheet).filter(
        Drascommentsheet.drasdocument_id == item.drasdocument_id
    ).all()
    for cs in cs_list:
        cs.current = False
        cs.last_stage = False
        ##cs.changed_by_fk = '1'
    db_session.commit() 
    
    cs_current = db_session.query(Drascommentsheet).filter(
        Drascommentsheet.drasdocument_id == item.drasdocument_id
    ).order_by(Drascommentsheet.actualDate.desc()).first()
    cs_current.current = True
    ##cs_current.changed_by_fk = '1'

    revisions = db_session.query(Drasrevision).filter(
        Drasrevision.drasdocument_id == item.drasdocument_id
    ).all()
    for rev in revisions:
        cs_last = db_session.query(Drascommentsheet).filter(
            Drascommentsheet.drasrevision_id == rev.id
        ).order_by(Drascommentsheet.actualDate.desc()).first()
        if cs_last:
            cs_last.last_stage = True
        ##cs_last.changed_by_fk = '1'
        db_session.commit()  
    db_session.commit() 


def update_all_current_and_last():
    db_session = db.session
    documents = db_session.query(Drasdocument).filter(
        Drasdocument.id == 370
    ).all()
    all_documents_len = len(documents)
    elaborate_docs = 0
    for doc in documents:
        elaborate_docs += 1
        cs_list = db_session.query(Drascommentsheet).filter(
            Drascommentsheet.drasdocument_id == doc.id
        ).all()
        for cs in cs_list:
            cs.current = False
            cs.last_stage = False
            cs.changed_by_fk = '1'
        db_session.commit() 
        
        cs_current = db_session.query(Drascommentsheet).filter(
            Drascommentsheet.drasdocument_id == doc.id
        ).order_by(Drascommentsheet.actualDate.desc()).first()
        cs_current.current = True
        cs_current.changed_by_fk = '1'

        revisions = db_session.query(Drasrevision).filter(
            Drasrevision.drasdocument_id == doc.id
        ).all()
        for rev in revisions:
            cs_last = db_session.query(Drascommentsheet).filter(
                Drascommentsheet.drasrevision_id == rev.id
            ).order_by(Drascommentsheet.actualDate.desc()).first()
            if cs_last:
                cs_last.last_stage = True
                cs_last.changed_by_fk = '1'
            db_session.commit()
        print(elaborate_docs,'/',all_documents_len, doc.name)  
    db_session.commit() 

def text_decode(comment):
    ''' Decode ASCII in comments'''
    if comment:
        string_encode = str(comment).encode("ascii", "ignore")
        string_decode = string_encode.decode()
        
        return string_decode
    return comment

def upload_all_comments():
    print('- - - - - - - - upload_all_comments - - START')
    db_session = db.session
    '''
    all_cs = db_session.query(Drascommentsheet).filter(
        Drascommentsheet.comments_uploaded == False,
        Drascommentsheet.stage != 'S',
    ).all()
    '''
    all_cs = db_session.query(Drascommentsheet).filter(
        #Drascommentsheet.comments_uploaded == False,
        Drascommentsheet.stage == 'YF',
    ).all()
    all_cs_len = len(all_cs)
    current_cs = 0
    #exceptions = {}
    for item in all_cs:
        current_cs += 1
        #print(current_cs,'/',all_cs_len,'Item:',item)
        comments = db_session.query(Drascomment).filter(
            Drascomment.drascommentsheet_id == item.id,
        ).all()
        comments_count = 0
        if comments:
            comments_count += len(comments)
            print('The item', item,'has comments!!!')
            item.comments_uploaded = True
            item.changed_by_fk = '1'
            db_session.commit()

            continue
        
        try:
            csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item.cs_file, data_only=True)
            csSheet = csFile.active
            #none_count = 0
            for row in csSheet.iter_rows(min_row=17,min_col=2):
                #print('CommentStatus', row[0].value,row[9].value,row[10].value,row[11].value, type(row[11].value))
                
                if row[0].value is not None and row[1].value is not None:
                    #print(row[0].value,'changed by', item.changed_by_fk)
                    comments_count += 1
                    if csSheet['L14'].value == 'Final Agreement':
                        comment = Drascomment(
                            drasdocument_id = item.drasdocument_id,
                            drasrevision_id = item.drasrevision_id,
                            drascommentsheet = item,
                            tagdiscipline= db_session.query(Tagdiscipline).filter(
                                                        Tagdiscipline.start <= int(row[1].value), 
                                                        Tagdiscipline.finish >= int(row[1].value)).first(), 

                            pos = str(row[0].value),
                            tag = row[1].value,
                            info = row[2].value,
                            ownerCommentBy = row[3].value,
                            ownerCommentDate = date_parse(csSheet['F15'].value),
                            ownerCommentComment = text_decode(row[4].value),

                            contractorReplyDate = date_parse(csSheet['H15'].value),
                            contractorReplyStatus = row[5].value,
                            contractorReplyComment = text_decode(row[6].value),
                            
                            ownerCounterReplyDate = date_parse(csSheet['K15'].value),
                            ownerCounterReplyComment = text_decode(row[8].value),

                            finalAgreementDate = date_parse(csSheet['M15'].value),
                            finalAgreemntCommentDate = date_parse(row[10].value),
                            finalAgreementComment = text_decode(row[11].value),

                            commentStatus = str(row[12].value),
                            
                        )      
                        #print('Contractor Status:',len(comment.contractorReplyStatus),comment.contractorReplyStatus)
                        #print('comment- created,changed',item.created_by_fk,item.changed_by_fk)
                        #comment.created_by_fk = item.created_by_fk
                        #comment.changed_by_fk = item.changed_by_fk
                        #db_session.add(comment)
                        #print('item,comment')
                        #print(item,comment)
                    else:
                        comment = Drascomment(
                            drasdocument_id = item.drasdocument_id,
                            drasrevision_id = item.drasrevision_id,
                            drascommentsheet = item,
                            tagdiscipline= db_session.query(Tagdiscipline).filter(
                                                        Tagdiscipline.start <= int(row[1].value), 
                                                        Tagdiscipline.finish >= int(row[1].value)).first(), 

                            pos = str(row[0].value),
                            tag = row[1].value,
                            info = row[2].value,
                            ownerCommentBy = row[3].value[:50],
                            ownerCommentDate = date_parse(csSheet['F15'].value),
                            ownerCommentComment = text_decode(row[4].value),

                            contractorReplyDate = date_parse(csSheet['H15'].value),
                            contractorReplyStatus = row[5].value,
                            contractorReplyComment = text_decode(row[6].value),
                            
                            ownerCounterReplyDate = date_parse(csSheet['J15'].value),
                            ownerCounterReplyComment = text_decode(row[7].value),

                            finalAgreementDate = date_parse(csSheet['L15'].value),
                            finalAgreemntCommentDate = date_parse(row[9].value),
                            finalAgreementComment = text_decode(row[10].value),

                            commentStatus = str(row[11].value),
                            
                        )
                    ## SET MAX LENGHT
                    if comment.commentStatus:
                        comment.commentStatus = comment.commentStatus[:20]
                    if comment.ownerCommentBy:
                        comment.ownerCommentBy = comment.ownerCommentBy[:50]
                    
                    if comment.pos:
                        comment.pos = str(comment.pos)[:5]

                    if comment.contractorReplyStatus:
                        comment.contractorReplyStatus = comment.contractorReplyStatus[:100]

                    #print('Contractor Status:',len(comment.contractorReplyStatus),comment.contractorReplyStatus)
                    #print('comment- created,changed',item.created_by_fk,item.changed_by_fk)
                    comment.created_by_fk = item.created_by_fk
                    comment.changed_by_fk = item.changed_by_fk
                    db_session.add(comment)
                    #print('item,comment')
                    #print(item,comment)
                
                '''    
                else:
                    
                    # Check on inifinite excel issue
                    none_count += 1
                    if none_count > 25:
                        #flash('Excel BAD FORMAT: Infinite Comments', category='info')
                        #raise Exception('Excel BAD FORMAT: Infinite Comments')
                        print(none_count,'--> INFINTE COMMENTS ERROR')

                '''
            item.comments_uploaded = True 
            #print('************ 12 **************',item.changed_by_fk)
            item.changed_by_fk = '1'   
            #session.query(Comment).filter(Comment.document_id == doc.id).delete() 
            #db_session.commit()  
            print(current_cs,'/',all_cs_len,'Item:',item,item.documentReferenceDoc,item.documentReferenceRev,'Comments:' ,comments_count)
            db_session.commit() 
        except Exception as e:
            print('* EXCEPTION * ', e)
            #exceptions[item.id] = e 
            db_session.rollback()
            #print('************ 13,5 **************')
            
            #flash('COMMENTS ERROR 003 | Non è stato possibile caricare i commenti per questo DRAS', category='warning')
            error_encode = str(e).encode("ascii", "ignore")
            string_decode = error_encode.decode()
            item.note = str(string_decode)

            item.comments_uploaded = False
            item.changed_by_fk = '1'
            #db_session.commit() 
            
        
        #print('************ 14 **************',item.changed_by_fk)
        #item.changed_by_fk = '1'
        db_session.commit()
    '''
    print('**************************')
    print('EXCEPTIONS')
    for k,v in exceptions.items():
        print(k,v)

    '''

def update_all_eng_data_from_cs(item):
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    
    
    csFile = openpyxl.load_workbook(UPLOAD_FOLDER+item.cs_file, data_only=True)
    csSheet = csFile.active
    
    deleted = session.query(Drascomment).filter(Drascomment.drascommentsheet_id == item.id).delete()
     
    #print('Delete Comments -+-+-+',item,deleted) 

    try:
        none_count = 0
        comments = 0
        
        for row in csSheet.iter_rows(min_row=17,min_col=2):
            
            
            if row[0].value is not None and row[1].value is not None:
                #print('pos',row[0].value,'tag',row[1].value,'info',row[2].value)
                
                if csSheet['L14'].value == 'Final Agreement':
                    comment = Drascomment(
                        drasdocument = item.drasdocument,
                        drasrevision = item.drasrevision,
                        drascommentsheet = item,
                        
                        tagdiscipline= session.query(Tagdiscipline).filter(
                                                    Tagdiscipline.start <= int(row[1].value), 
                                                    Tagdiscipline.finish >= int(row[1].value)).first(), 
                    

                        pos = row[0].value,
                        tag = row[1].value,
                        info = row[2].value,
                        ownerCommentBy = row[3].value,
                        ownerCommentDate = date_parse(csSheet['F15'].value),
                        ownerCommentComment = text_decode(row[4].value) ,

                        contractorReplyDate = date_parse(csSheet['H15'].value),
                        contractorReplyStatus = row[5].value,
                        contractorReplyComment = text_decode(row[6].value),
                        
                        ownerCounterReplyDate = date_parse(csSheet['K15'].value),
                        ownerCounterReplyComment = text_decode(row[8].value),

                        finalAgreementDate = date_parse(csSheet['M15'].value),
                        finalAgreemntCommentDate = date_parse(row[10].value),
                        finalAgreementComment = text_decode(row[11].value),

                        commentStatus = str(row[12].value),
                    )

                    #if item.current:
                        #comment.drasdocument_id = doc.id
                    #print('Contractor Status:',len(comment.contractorReplyStatus),comment.contractorReplyStatus)
                    #session.add(comment)
                    #session.commit()
                else:
                    comment = Drascomment(
                        drasdocument = item.drasdocument,
                        drasrevision = item.drasrevision,
                        drascommentsheet = item,
                        
                        tagdiscipline= session.query(Tagdiscipline).filter(
                                                    Tagdiscipline.start <= int(row[1].value), 
                                                    Tagdiscipline.finish >= int(row[1].value)).first(), 

                        pos = row[0].value,
                        tag = row[1].value,
                        info = row[2].value,
                        
                        ownerCommentBy = row[3].value,
                        ownerCommentDate = date_parse(csSheet['F15'].value),
                        ownerCommentComment = text_decode(row[4].value),

                        contractorReplyDate = date_parse(csSheet['H15'].value),
                        contractorReplyStatus = row[5].value,
                        contractorReplyComment = text_decode(row[6].value),
                        
                        ownerCounterReplyDate = date_parse(csSheet['J15'].value),
                        ownerCounterReplyComment = text_decode(row[7].value),

                        finalAgreementDate = date_parse(csSheet['L15'].value),
                        finalAgreemntCommentDate = date_parse(row[9].value),
                        finalAgreementComment = text_decode(row[10].value),

                        commentStatus = str(row[11].value),
                        
                    )
                ## SET MAX LENGHT
                if comment.commentStatus:
                    comment.commentStatus = comment.commentStatus[:20]
                if comment.ownerCommentBy:
                    comment.ownerCommentBy = comment.ownerCommentBy[:50]
                
                if comment.pos:
                    comment.pos = str(comment.pos)[:5]

                if comment.contractorReplyStatus:
                    comment.contractorReplyStatus = comment.contractorReplyStatus[:100]

                comment.created_by_fk = '1'
                comment.changed_by_fk = '1'
                
                comments += 1
                session.add(comment)

            else:
                # Check on inifinite excel issue
                none_count += 1
                if none_count > 35:
                    break
                    #raise Exception('Excel BAD FORMAT: Infinite Comments')
        #session.query(Comment).filter(Comment.document_id == doc.id).delete()
        session.commit()
        
        return item, comments 
    except Exception as e:
        session.rollback()


def piero_items():
    #item ='4def885a-604b-11e9-bffd-ac87a32187da_sep_DRAS_2544-17-DW-0510-04_CY.xlsx'
    session = db.session
    csFile = openpyxl.load_workbook('app/comments/Piero_eng.xlsx', data_only=True)
    csSheet = csFile.active

    start = 0 

    for row in csSheet.iter_rows(min_row=2, min_col=1):
        if row[2].value is None:
            start += 1
            try:
                document = row[0].value[:str(row[0].value).rindex('_')]
                rev = row[0].value[str(row[0].value).rindex('_')+1:]
                
                dras_doc = session.query(Drasdocument).filter(Drasdocument.name == document).first()
                
                if dras_doc: 
                    dras_rev = session.query(Drasrevision).filter(
                        Drasrevision.name == rev,
                        Drasrevision.drasdocument_id == dras_doc.id).first()

                    if dras_rev:
                        dras = session.query(Drascommentsheet).filter(
                            Drascommentsheet.stage == 'YF',
                            Drascommentsheet.drasrevision_id == dras_rev.id).first() 
                        
                        if dras:
                            dras, comments = update_all_eng_data_from_cs(dras)
                            dras.note = 'Piero003'
                            dras.changed_by_fk = '1'
                            session.commit()
                            
                            print(start,dras.drasdocument, dras.drasrevision, dras.stage, 'C: ',comments)

                            csSheet.cell(row=row[0].row,column=3).value = 'OK'
                            csSheet.cell(row=row[0].row,column=4).value = comments 
                        else:
                            print(start,' ',document,'DRAS | NOT FOUND')
                            csSheet.cell(row=row[0].row,column=3).value = 'NO DRAS'
                    else:
                            csSheet.cell(row=row[0].row,column=3).value = 'NO REV'
                            print(start,' ',document,'REV | NOT FOUND')
                else:
                            csSheet.cell(row=row[0].row,column=3).value = 'NO DOC' 
                            print(start,' ',document,'DOC | NOT FOUND') 
                
                
            except Exception as e:
                csSheet.cell(row=row[0].row,column=3).value = 'Error'
                csSheet.cell(row=row[0].row,column=5).value = str(e) 
                print(e) 
    
    csFile.save('app/comments/Piero_eng.xlsx')
    csFile.close() 
